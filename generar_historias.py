"""
Genera el Excel de Historias Épicas y de Usuario para el proyecto !Blood.
Estructura idéntica a los ejemplos de las imágenes de referencia.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────
# PALETA DE COLORES
# ──────────────────────────────────────────────────────────────
DARK_HEADER   = "2E3440"   # gris oscuro para encabezados principales
BLUE_HEADER   = "1F3864"   # azul oscuro para sub-encabezados
MID_BLUE      = "2F5496"   # azul medio para enunciado
LIGHT_BLUE    = "BDD7EE"   # azul claro para sub-encabezado criterios
ORANGE_LINK   = "C55A11"   # naranja/teja para texto de "Quiero" e IDs
BLACK         = "000000"
WHITE         = "FFFFFF"
ROW_ALT       = "F2F2F2"   # filas alternadas claras

THIN  = Side(style="thin",   color=BLACK)
MED   = Side(style="medium", color=BLACK)

def border_all(thick=False):
    s = MED if thick else THIN
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=BLACK, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def center_wrap():
    return Alignment(horizontal="center", vertical="center",
                     wrap_text=True)

def left_wrap():
    return Alignment(horizontal="left", vertical="center",
                     wrap_text=True)

# ──────────────────────────────────────────────────────────────
# DATOS: HISTORIAS ÉPICAS
# ──────────────────────────────────────────────────────────────
epicas = [
    ("HE-01", "Gestión de Acceso de Usuarios.",
     "El sistema debe permitir gestionar el acceso de los usuarios a la plataforma !Blood "
     "(registro en tres pasos con selector de tipo de documento, fotos de documento + selfie con cámara, "
     "inicio de sesión y recuperación de contraseña)."),

    ("HE-02", "Gestión del Perfil de Usuario.",
     "El sistema debe permitir a cada usuario consultar y actualizar su información personal, "
     "tipo de sangre, calificación por estrellas, estadísticas de donación, encuesta de aptitud "
     "y configuración de cuenta."),

    ("HE-03", "Gestión de Solicitudes de Sangre.",
     "El sistema debe permitir crear, publicar, explorar y gestionar solicitudes de donación de "
     "sangre con filtros por compatibilidad y cercanía geográfica. El detalle de la solicitud "
     "muestra el botón 'Quiero Donar' siempre visible sin necesidad de desplazamiento."),

    ("HE-04", "Gestión del Proceso de Donación.",
     "El sistema debe guiar al donante en la confirmación de donación (checklist de salud) e "
     "informar que la cancelación posterior implica una penalización de 30 días sin poder donar."),

    ("HE-05", "Sistema de Calificación Mutua.",
     "El sistema debe permitir que, tras completarse una donación, tanto el donante como el "
     "solicitante se califiquen mutuamente con 1-5 estrellas, un aviso de equidad y un comentario "
     "opcional, contribuyendo a la reputación visible en el perfil."),

    ("HE-06", "Gestión de Campañas de Donación.",
     "El sistema debe permitir a los profesionales de salud crear, publicar y gestionar campañas "
     "institucionales de donación (con mensaje motivacional, sin campo de teléfono) y a todos los "
     "usuarios explorarlas y reservar cupos. Incluye las vistas exclusivas del profesional: "
     "Home Pro, Mis Campañas, Gestión de Campaña y Perfil Pro."),

    ("HE-07", "Sistema de Notificaciones.",
     "El sistema debe notificar en tiempo real sobre solicitudes urgentes compatibles, "
     "aceptación de donación por parte de un donante, recordatorio de campaña próxima y "
     "reconocimientos de logros. Se eliminó la notificación de 'nuevo mensaje' por supresión del chat."),

    ("HE-08", "Moderación y Seguridad de Contenido.",
     "El sistema debe garantizar la privacidad de los datos de los usuarios, gestionar el "
     "proceso de verificación de identidad (foto frontal + trasera del documento + selfie) y "
     "prevenir el uso indebido de la plataforma."),

    ("HE-09", "Estadísticas y Gamificación.",
     "El sistema debe mostrar estadísticas individuales y comunitarias, otorgar insignias, "
     "mostrar la calificación promedio por estrellas en el perfil y enviar mensajes motivacionales "
     "para fomentar la donación recurrente."),

    ("HE-10", "Compatibilidad Sanguínea y Geolocalización.",
     "El sistema debe determinar automáticamente la compatibilidad entre tipos de sangre y "
     "calcular la distancia entre el usuario y cada solicitud o campaña."),

    ("HE-11", "Sistema de Penalizaciones.",
     "El sistema debe aplicar penalizaciones automáticas por: (1) abuso de urgencia en solicitudes "
     "→ bloqueo temporal de creación de solicitudes; (2) cancelación de donación confirmada "
     "→ 30 días sin poder confirmar donaciones; (3) calificación discriminatoria o falsa "
     "→ suspensión de cuenta."),

    ("HE-12", "Encuesta de Aptitud para Donación.",
     "El sistema debe presentar al usuario una encuesta de 8 preguntas de sí/no tras completar "
     "el registro (Paso 3), determinar su aptitud para donar y permitirle retomar la encuesta "
     "desde su perfil en cualquier momento."),
]

# ──────────────────────────────────────────────────────────────
# DATOS: HISTORIAS DE USUARIO  (por hoja / épica)
# ──────────────────────────────────────────────────────────────
# Estructura de cada escenario:
#  (num_escenario, titulo_criterio, contexto, evento, resultado_esperado)
# Estructura de cada HU:
#  (id_epica, id_historia, rol, quiero, para, [escenarios])

hu_por_epica = {

# ══════════════════════════════════════════════════════════════
"HE-01": [
  ("HE-01","HU-1","Yo como USUARIO",
   "Necesito iniciar sesión en la aplicación.",
   "Para poder hacer uso de todas las funcionalidades de !Blood.",
   [
     (1,"Inicio de sesión exitoso.",
      "En caso de ingresar correo/usuario y contraseña correctos.",
      "Clic en el botón INGRESAR.",
      "El sistema redirige a la pantalla principal (Home)."),
     (2,"Inicio de sesión fallido.",
      "En caso de ingresar correo/usuario o contraseña incorrectos.",
      "Clic en el botón INGRESAR.",
      "El sistema muestra el mensaje: 'Usuario o contraseña incorrectos'."),
     (3,"Múltiples intentos fallidos.",
      "En caso de ingresar credenciales incorrectas cinco veces consecutivas.",
      "Clic en el botón INGRESAR.",
      "El sistema bloquea el acceso temporalmente y muestra el mensaje: 'Demasiados intentos fallidos. Intente de nuevo en 15 minutos'."),
     (4,"Campos obligatorios vacíos.",
      "En caso de no ingresar correo/usuario o contraseña.",
      "Clic en el botón INGRESAR.",
      "El sistema muestra el mensaje: 'Todos los campos son obligatorios', resaltando los campos vacíos."),
     (5,"Redirección no autorizada.",
      "En caso de intentar acceder a una pantalla protegida sin haber iniciado sesión.",
      "Ingreso de la ruta en la barra de navegación.",
      "El sistema redirige al usuario a la pantalla de Inicio de Sesión."),
   ]
  ),
  ("HE-01","HU-1.1","Yo como USUARIO",
   "Necesito recuperar mi contraseña.",
   "Para poder recuperar el acceso a mi cuenta en caso de olvidarla.",
   [
     (1,"Recuperación exitosa.",
      "En caso de ingresar un correo registrado en el sistema.",
      "Clic en el botón RECUPERAR.",
      "El sistema muestra el mensaje: 'Se ha enviado un correo con instrucciones para recuperar tu contraseña'."),
     (2,"Correo no reconocido.",
      "En caso de ingresar un correo no registrado en el sistema.",
      "Clic en el botón RECUPERAR.",
      "El sistema muestra el mensaje: 'No encontramos una cuenta asociada a ese correo'."),
     (3,"Campo vacío.",
      "En caso de no ingresar ningún correo.",
      "Clic en el botón RECUPERAR.",
      "El sistema muestra el mensaje: 'El correo electrónico es obligatorio'."),
     (4,"Regreso a inicio de sesión.",
      "En caso de querer regresar sin recuperar la contraseña.",
      "Clic en el enlace o botón VOLVER.",
      "El sistema regresa a la pantalla de Inicio de Sesión (HU-1)."),
   ]
  ),
  ("HE-01","HU-1.2","Yo como USUARIO NUEVO",
   "Necesito registrarme en la aplicación en tres pasos.",
   "Para poder acceder a las funcionalidades de !Blood como donante o solicitante.",
   [
     (1,"Registro exitoso – Paso 1 (Datos personales).",
      "En caso de ingresar nombre completo, fecha de nacimiento, tipo de sangre, número de documento, "
      "tipo de documento (TI / CC / CE / PA / NIT) y tipo de usuario (Ciudadano / Profesional de salud). "
      "Nota: el Paso 1 ya NO solicita teléfono ni correo.",
      "Clic en el botón SIGUIENTE.",
      "El sistema avanza al Paso 2 y muestra el indicador de progreso en la segunda etapa."),
     (2,"Registro fallido – Paso 1, campos inválidos.",
      "En caso de no seleccionar el tipo de documento o ingresar un número de documento con menos de 6 caracteres.",
      "Clic en el botón SIGUIENTE.",
      "El sistema muestra mensajes de validación resaltando los campos inválidos."),
     (3,"Registro exitoso – Paso 2 (Verificación de identidad con documento y selfie).",
      "En caso de cargar la foto frontal del documento, la foto trasera del documento y tomar la selfie "
      "con cámara obligatoriamente (no se permite cargar desde galería para la selfie), más ciudad y "
      "dirección (o usar geolocalización).",
      "Clic en el botón SIGUIENTE.",
      "El sistema valida las tres imágenes y avanza al Paso 3."),
     (4,"Selfie desde galería bloqueada.",
      "En caso de intentar cargar la selfie desde la galería del dispositivo en lugar de usar la cámara.",
      "El usuario selecciona la opción de galería en el campo de selfie.",
      "El sistema muestra el mensaje: 'La selfie debe tomarse con la cámara del dispositivo por razones de seguridad' y mantiene el campo vacío."),
     (5,"Registro exitoso – Paso 3 (Configuración de cuenta).",
      "En caso de ingresar nombre de usuario único, contraseña segura, confirmar contraseña, "
      "correo electrónico y aceptar términos y condiciones.",
      "Clic en el botón REGISTRARME.",
      "El sistema crea la cuenta, muestra pantalla de bienvenida y redirige automáticamente a la encuesta de aptitud (HE-12)."),
     (6,"Contraseña débil.",
      "En caso de ingresar una contraseña con fortaleza baja (menos de 8 caracteres o sin combinación de tipos).",
      "El usuario escribe en el campo contraseña.",
      "El sistema muestra el indicador de fortaleza en 'Baja' con sugerencias de mejora."),
     (7,"Nombre de usuario ya existente.",
      "En caso de ingresar un nombre de usuario que ya está registrado.",
      "Clic en el botón REGISTRARME.",
      "El sistema muestra el mensaje: 'Ese nombre de usuario ya está en uso. Elige otro'."),
   ]
  ),
],

# ══════════════════════════════════════════════════════════════
"HE-02": [
  ("HE-02","HU-2","Yo como USUARIO AUTENTICADO",
   "Necesito ver y editar mi perfil.",
   "Para mantener mi información personal y tipo de sangre actualizados.",
   [
     (1,"Ver perfil completo.",
      "En caso de haber iniciado sesión y navegar a la sección Perfil.",
      "Clic en el ícono PERFIL en la barra de navegación.",
      "El sistema muestra: foto, nombre, tipo de sangre, calificación por estrellas (promedio), insignias, "
      "estadísticas (donaciones, vidas ayudadas, solicitudes) y opciones: Encuesta de Aptitud, Privacidad, "
      "Términos, Cerrar Sesión. Nota: no se muestran teléfono ni correo en la vista pública del perfil."),
     (2,"Editar foto de perfil.",
      "En caso de estar en la pantalla de perfil.",
      "Clic en el botón EDITAR FOTO.",
      "El sistema abre el selector de imagen y actualiza la foto al guardar."),
     (3,"Editar datos personales.",
      "En caso de estar en la pantalla de edición de perfil.",
      "Modificar campos y clic en GUARDAR.",
      "El sistema actualiza los datos y muestra el mensaje: 'Perfil actualizado correctamente'."),
     (4,"Acceder a encuesta de aptitud desde perfil.",
      "En caso de querer revisar o retomar la encuesta de aptitud para donación.",
      "Clic en la opción ENCUESTA DE APTITUD en el menú del perfil.",
      "El sistema redirige a la pantalla de Encuesta de Aptitud (HE-12) mostrando el estado actual."),
     (5,"Cerrar sesión.",
      "En caso de estar en la pantalla de perfil.",
      "Clic en CERRAR SESIÓN.",
      "El sistema cierra la sesión y redirige a la pantalla de Inicio de Sesión."),
   ]
  ),
  ("HE-02","HU-2.1","Yo como USUARIO AUTENTICADO",
   "Necesito ver mis estadísticas de donación y mi calificación.",
   "Para conocer mi historial, impacto dentro de la comunidad !Blood y reputación como donante.",
   [
     (1,"Ver estadísticas en perfil.",
      "En caso de haber realizado al menos una donación.",
      "Navegar a la sección PERFIL.",
      "El sistema muestra: número de donaciones, vidas ayudadas, solicitudes creadas y calificación promedio en estrellas."),
     (2,"Ver estadísticas en Home.",
      "En caso de estar en la pantalla principal.",
      "El usuario visualiza la tarjeta de sangre en el Home.",
      "El sistema muestra tipo de sangre, donaciones realizadas y vidas ayudadas actualizadas."),
     (3,"Sin calificaciones aún.",
      "En caso de que el usuario no haya recibido ninguna calificación todavía.",
      "Navegar a la sección PERFIL.",
      "El sistema muestra las estrellas vacías o en gris con el texto: 'Aún sin calificaciones'."),
   ]
  ),
],

# ══════════════════════════════════════════════════════════════
"HE-03": [
  ("HE-03","HU-3","Yo como USUARIO AUTENTICADO",
   "Necesito crear una solicitud de donación de sangre.",
   "Para notificar a donantes compatibles y cercanos sobre la necesidad urgente.",
   [
     (1,"Creación exitosa.",
      "En caso de seleccionar tipo de sangre, unidades (1-10), nivel de urgencia y centro de salud.",
      "Clic en el botón PUBLICAR SOLICITUD.",
      "El sistema publica la solicitud y muestra el mensaje: 'Tu solicitud fue publicada. X donantes compatibles han sido notificados'."),
     (2,"Campos obligatorios vacíos.",
      "En caso de no seleccionar tipo de sangre, urgencia o centro de salud.",
      "Clic en el botón PUBLICAR SOLICITUD.",
      "El sistema resalta los campos faltantes y muestra el mensaje: 'Completa los campos obligatorios'."),
     (3,"Solicitud con mensaje opcional.",
      "En caso de ingresar un mensaje para los donantes (máx. 200 caracteres).",
      "Clic en el botón PUBLICAR SOLICITUD.",
      "El sistema incluye el mensaje en la tarjeta de solicitud visible para donantes compatibles."),
     (4,"Límite de caracteres del mensaje.",
      "En caso de escribir más de 200 caracteres en el campo de mensaje.",
      "El usuario continúa escribiendo.",
      "El sistema bloquea el ingreso de más caracteres y muestra el contador: '200/200'."),
     (5,"Selector de urgencia – advertencia por abuso.",
      "En caso de que el sistema detecte que el usuario ha creado más de 3 solicitudes con urgencia 'Crítica' "
      "en los últimos 7 días sin que ninguna haya sido completada.",
      "El usuario selecciona el nivel de urgencia 'Crítica'.",
      "El sistema muestra el modal de advertencia: 'El uso indebido de la urgencia puede resultar en la "
      "restricción temporal de tu cuenta (HE-11)' y solicita confirmación antes de publicar."),
   ]
  ),
  ("HE-03","HU-3.1","Yo como USUARIO AUTENTICADO",
   "Necesito explorar el listado de solicitudes activas.",
   "Para identificar solicitudes de sangre compatibles con mi tipo y cercanas a mi ubicación.",
   [
     (1,"Ver listado de solicitudes.",
      "En caso de haber iniciado sesión y navegar a la pantalla de solicitudes.",
      "Clic en VER SOLICITUDES o en el ícono EXPLORAR.",
      "El sistema muestra la lista de solicitudes activas con: urgencia, tipo de sangre, solicitante, centro, distancia y tiempo de publicación."),
     (2,"Filtrar por compatibilidad.",
      "En caso de seleccionar el filtro 'Compatible'.",
      "Clic en el chip COMPATIBLE.",
      "El sistema muestra únicamente las solicitudes cuyo tipo de sangre es compatible con el del usuario autenticado."),
     (3,"Filtrar por urgencias.",
      "En caso de seleccionar el filtro 'Urgentes'.",
      "Clic en el chip URGENTES.",
      "El sistema muestra únicamente las solicitudes con nivel de urgencia 'Urgente'."),
     (4,"Buscar por nombre o tipo de sangre.",
      "En caso de ingresar texto en la barra de búsqueda.",
      "El usuario escribe en el campo de búsqueda.",
      "El sistema filtra en tiempo real las solicitudes que coincidan con el texto ingresado."),
     (5,"Sin solicitudes en la zona.",
      "En caso de no existir solicitudes activas en la zona del usuario.",
      "El usuario accede a la pantalla de solicitudes.",
      "El sistema muestra el mensaje: 'No hay solicitudes activas en tu zona por el momento'."),
   ]
  ),
  ("HE-03","HU-3.2","Yo como USUARIO AUTENTICADO",
   "Necesito ver el detalle de una solicitud de sangre.",
   "Para decidir si puedo y quiero donar según la compatibilidad y ubicación.",
   [
     (1,"Ver detalle completo con botón visible.",
      "En caso de seleccionar una solicitud del listado.",
      "Clic en la tarjeta de solicitud.",
      "El sistema muestra en una sola pantalla (sin necesidad de scroll): urgencia, tipo de sangre, unidades, "
      "datos del solicitante verificado, mensaje, centro de salud, distancia, estado de donantes, aviso de "
      "compatibilidad y el botón QUIERO DONAR siempre visible al fondo de la pantalla."),
     (2,"Compatibilidad automática detectada.",
      "En caso de que el tipo de sangre del usuario sea compatible con el requerido.",
      "El usuario abre el detalle de la solicitud.",
      "El sistema muestra el banner: 'Tu tipo de sangre [X] es compatible con esta solicitud'."),
     (3,"Incompatibilidad detectada.",
      "En caso de que el tipo de sangre del usuario NO sea compatible.",
      "El usuario abre el detalle de la solicitud.",
      "El sistema muestra el aviso: 'Tu tipo de sangre no es compatible con esta solicitud' y deshabilita el botón QUIERO DONAR."),
   ]
  ),
  ("HE-03","HU-3.3","Yo como USUARIO AUTENTICADO",
   "Necesito gestionar mis solicitudes publicadas.",
   "Para dar seguimiento al estado de mis solicitudes y cancelarlas si ya no son necesarias.",
   [
     (1,"Ver solicitudes propias.",
      "En caso de navegar a la sección MIS SOLICITUDES.",
      "Clic en la opción MIS SOLICITUDES del menú.",
      "El sistema muestra las solicitudes organizadas en pestañas: Activas, Completadas y Canceladas."),
     (2,"Ver progreso de donantes.",
      "En caso de tener una solicitud activa con al menos un donante confirmado.",
      "El usuario visualiza la tarjeta de solicitud activa.",
      "El sistema muestra la barra de progreso con los avatares de los donantes confirmados sobre el total requerido."),
     (3,"Cancelar solicitud.",
      "En caso de querer cancelar una solicitud activa.",
      "Clic en el botón CANCELAR de la solicitud.",
      "El sistema solicita confirmación y, al confirmar, mueve la solicitud a la pestaña 'Canceladas' y notifica a los donantes que habían aceptado."),
   ]
  ),
],

# ══════════════════════════════════════════════════════════════
"HE-04": [
  ("HE-04","HU-4","Yo como DONANTE POTENCIAL",
   "Necesito confirmar mi disposición a donar sangre para una solicitud.",
   "Para que el solicitante sepa que acudiré al centro de salud indicado.",
   [
     (1,"Confirmación exitosa.",
      "En caso de revisar el resumen de compromiso, completar el checklist de salud y aceptar las condiciones.",
      "Clic en el botón CONFIRMAR MI DONACIÓN.",
      "El sistema registra la donación, muestra la pantalla de agradecimiento con el nombre del centro al que debe "
      "acudir y notifica al solicitante: '[Nombre] ha confirmado que donará [tipo] en [centro]'."),
     (2,"Aviso de penalización por cancelación.",
      "En caso de haber confirmado una donación y estar visualizando el resumen.",
      "El donante visualiza la pantalla de confirmación de donación.",
      "El sistema muestra el aviso: 'Importante: si cancelas esta donación después de confirmar, tu cuenta quedará "
      "suspendida 30 días para confirmar nuevas donaciones (HE-11)'."),
     (3,"Checklist incompleto.",
      "En caso de no marcar todos los ítems obligatorios del checklist de salud (edad 18-65, peso >50 kg, "
      "sin enfermedades activas, 3 meses desde última donación).",
      "Clic en el botón CONFIRMAR MI DONACIÓN.",
      "El sistema muestra el mensaje: 'Debes confirmar todos los requisitos de salud para continuar', resaltando los ítems pendientes."),
     (4,"Cancelar antes de confirmar.",
      "En caso de estar en la pantalla de confirmación y decidir no continuar.",
      "Clic en el botón CANCELAR o flecha de retroceso.",
      "El sistema regresa al detalle de la solicitud sin registrar ninguna acción ni penalización."),
     (5,"Donante ya registrado en esa solicitud.",
      "En caso de intentar confirmar una donación en una solicitud donde ya fue registrado como donante.",
      "El usuario accede al detalle de la solicitud.",
      "El sistema muestra el estado 'Ya confirmaste tu donación para esta solicitud' y deshabilita el botón QUIERO DONAR."),
   ]
  ),
],

# ══════════════════════════════════════════════════════════════
"HE-05": [
  ("HE-05","HU-5","Yo como USUARIO AUTENTICADO (DONANTE O SOLICITANTE)",
   "Necesito calificar a la otra parte tras completarse una donación.",
   "Para contribuir a la reputación de la comunidad y garantizar un entorno de confianza en !Blood.",
   [
     (1,"Calificación exitosa como donante.",
      "En caso de haberse completado una donación donde el usuario actuó como donante.",
      "El sistema activa la pantalla de calificación tras el registro de la donación completada.",
      "El sistema muestra la pantalla de Calificación al Solicitante con escala 1-5 estrellas, aviso de equidad "
      "('Tu calificación es anónima y contribuye a la reputación de la comunidad') y campo de comentario opcional."),
     (2,"Calificación exitosa como solicitante.",
      "En caso de haberse completado una donación donde el usuario actuó como solicitante.",
      "El sistema activa la pantalla de calificación al donante.",
      "El sistema muestra la pantalla de Calificación al Donante con escala 1-5 estrellas, aviso de equidad y campo de comentario opcional (máx. 200 caracteres)."),
     (3,"Calificación con comentario.",
      "En caso de escribir un comentario en el campo opcional.",
      "El usuario escribe el comentario y presiona ENVIAR CALIFICACIÓN.",
      "El sistema registra la calificación con el comentario y actualiza el promedio de estrellas en el perfil del calificado."),
     (4,"Calificación sin comentario.",
      "En caso de seleccionar estrellas pero dejar el campo de comentario vacío.",
      "Clic en el botón ENVIAR CALIFICACIÓN.",
      "El sistema registra la calificación numérica sin comentario y actualiza el promedio de estrellas en el perfil."),
     (5,"Calificación discriminatoria detectada.",
      "En caso de que el comentario contenga lenguaje discriminatorio, ofensivo o que infrinja las normas de la comunidad.",
      "Clic en ENVIAR CALIFICACIÓN.",
      "El sistema bloquea el envío, registra el intento y muestra el aviso: 'Tu comentario infringe las normas de la comunidad. "
      "Las calificaciones discriminatorias pueden resultar en la suspensión de tu cuenta (HE-11)'."),
   ]
  ),
],

# ══════════════════════════════════════════════════════════════
"HE-06": [
  ("HE-06","HU-6","Yo como PROFESIONAL DE SALUD",
   "Necesito crear y publicar una campaña de donación de sangre.",
   "Para organizar jornadas institucionales y alcanzar a la mayor cantidad de donantes potenciales.",
   [
     (1,"Creación exitosa de campaña.",
      "En caso de ingresar nombre, centro de salud, dirección, fecha, hora inicio, hora fin, cupos, "
      "tipos de sangre y descripción (máx. 500 caracteres). Nota: el formulario ya NO incluye campo de teléfono de contacto.",
      "Clic en el botón PUBLICAR CAMPAÑA.",
      "El sistema publica la campaña, muestra la pantalla de confirmación con un mensaje motivacional "
      "('¡Gracias por organizar esta jornada! Tu iniciativa puede salvar muchas vidas.'), el alcance estimado "
      "de donantes notificados, y la campaña aparece en el listado público."),
     (2,"Campaña con todos los tipos de sangre.",
      "En caso de seleccionar la opción 'Todos' en tipos de sangre necesarios.",
      "Clic en PUBLICAR CAMPAÑA.",
      "El sistema notifica a todos los usuarios con cualquier tipo de sangre en la zona de la campaña."),
     (3,"Campos obligatorios incompletos.",
      "En caso de no ingresar nombre, dirección, fecha u hora de inicio.",
      "Clic en PUBLICAR CAMPAÑA.",
      "El sistema resalta los campos faltantes y muestra el mensaje: 'Completa los campos obligatorios para publicar la campaña'."),
     (4,"Geolocalización de dirección.",
      "En caso de querer usar la ubicación actual como dirección del evento.",
      "Clic en el botón de geolocalización.",
      "El sistema captura la ubicación GPS del dispositivo y la rellena automáticamente en el campo dirección."),
   ]
  ),
  ("HE-06","HU-6.1","Yo como USUARIO AUTENTICADO",
   "Necesito explorar las campañas de donación disponibles.",
   "Para conocer las jornadas próximas y reservar un cupo para donar.",
   [
     (1,"Ver listado de campañas.",
      "En caso de navegar a la sección CAMPAÑAS.",
      "Clic en CAMPAÑAS desde el Home o el menú.",
      "El sistema muestra las campañas disponibles con: fecha, institución verificada, dirección, horario, distancia y disponibilidad de cupos."),
     (2,"Filtrar por proximidad.",
      "En caso de seleccionar el filtro 'Cercanas'.",
      "Clic en el chip CERCANAS.",
      "El sistema ordena las campañas de menor a mayor distancia respecto a la ubicación del usuario."),
     (3,"Ver detalle de campaña.",
      "En caso de seleccionar una campaña del listado.",
      "Clic en la tarjeta de campaña.",
      "El sistema muestra: organización verificada, descripción, mapa, horario, tipos de sangre (con indicador de más urgentes), requisitos, indicador de cupos y botón RESERVAR MI CUPO."),
     (4,"Reservar cupo.",
      "En caso de querer reservar un cupo en una campaña.",
      "Clic en el botón RESERVAR MI CUPO.",
      "El sistema registra la reserva, actualiza la barra de cupos y muestra el mensaje: 'Tu cupo ha sido reservado. Te esperamos el [fecha] en [lugar]'."),
     (5,"Campaña sin cupos disponibles.",
      "En caso de que todos los cupos de la campaña hayan sido reservados.",
      "El usuario accede al detalle de la campaña.",
      "El sistema muestra el estado 'Sin cupos disponibles' y deshabilita el botón RESERVAR MI CUPO."),
   ]
  ),
  ("HE-06","HU-6.2","Yo como PROFESIONAL DE SALUD",
   "Necesito ver el panel de inicio exclusivo para profesionales (Home Pro).",
   "Para tener visibilidad rápida del estado de mis campañas, donantes alcanzados y estadísticas de impacto.",
   [
     (1,"Ver Home Pro con estadísticas.",
      "En caso de iniciar sesión con una cuenta de tipo Profesional de Salud.",
      "El sistema carga la pantalla principal.",
      "El sistema muestra el Home Pro con: saludo personalizado, total de campañas activas, total de donantes alcanzados, "
      "total de cupos reservados, tarjetas de campañas recientes y acceso directo a Mis Campañas."),
     (2,"Acceso rápido a gestión de campaña.",
      "En caso de estar en el Home Pro y seleccionar una campaña reciente.",
      "Clic en la tarjeta de campaña.",
      "El sistema navega directamente a la pantalla de Gestión de Campaña (HU-6.3)."),
     (3,"Sin campañas activas.",
      "En caso de que el profesional aún no haya creado ninguna campaña.",
      "El profesional accede al Home Pro.",
      "El sistema muestra el mensaje: 'No tienes campañas activas. Crea tu primera campaña y comienza a salvar vidas.' con el botón CREAR CAMPAÑA."),
   ]
  ),
  ("HE-06","HU-6.3","Yo como PROFESIONAL DE SALUD",
   "Necesito gestionar el listado de mis campañas publicadas (Mis Campañas).",
   "Para hacer seguimiento, editar o cancelar campañas desde una pantalla centralizada.",
   [
     (1,"Ver listado de mis campañas.",
      "En caso de navegar a la sección MIS CAMPAÑAS desde el Home Pro o el menú.",
      "Clic en MIS CAMPAÑAS.",
      "El sistema muestra todas las campañas del profesional organizadas en pestañas: Activas, Finalizadas y Canceladas, "
      "con datos de: nombre, fecha, cupos reservados/totales y estado."),
     (2,"Acceder a gestión detallada.",
      "En caso de seleccionar una campaña del listado.",
      "Clic en la tarjeta de campaña activa.",
      "El sistema navega a la pantalla de Gestión de Campaña con el panel completo de administración."),
     (3,"Cancelar campaña.",
      "En caso de querer cancelar una campaña activa.",
      "Clic en la opción CANCELAR de la tarjeta de campaña.",
      "El sistema solicita confirmación; al confirmar, cancela la campaña, notifica a los donantes con cupo reservado y mueve la campaña a la pestaña 'Canceladas'."),
   ]
  ),
  ("HE-06","HU-6.4","Yo como PROFESIONAL DE SALUD",
   "Necesito gestionar el detalle de una campaña activa (Gestión Campaña).",
   "Para administrar los cupos, visualizar los donantes registrados y monitorear el avance en tiempo real.",
   [
     (1,"Ver panel de gestión completo.",
      "En caso de acceder al detalle desde Mis Campañas.",
      "Clic en la tarjeta de campaña activa.",
      "El sistema muestra: nombre, fecha/hora, dirección, barra de cupos (reservados/totales), lista de donantes "
      "registrados con tipo de sangre y estado, y opciones de editar o cancelar campaña."),
     (2,"Editar campaña activa.",
      "En caso de querer modificar datos de una campaña publicada.",
      "Clic en el botón EDITAR.",
      "El sistema abre el formulario de edición con los campos actuales precargados. Cambios guardados actualizan la vista pública instantáneamente."),
     (3,"Ver lista de donantes registrados.",
      "En caso de tener al menos un donante con cupo reservado.",
      "El profesional visualiza la pantalla de Gestión Campaña.",
      "El sistema muestra la lista de donantes con: nombre, tipo de sangre, hora de reserva y estado (confirmado / pendiente / asistió)."),
   ]
  ),
],

# ══════════════════════════════════════════════════════════════
"HE-07": [
  ("HE-07","HU-7","Yo como USUARIO AUTENTICADO",
   "Necesito recibir notificaciones relevantes en tiempo real.",
   "Para estar al tanto de solicitudes urgentes compatibles, estados de mis donaciones y actividad de la comunidad.",
   [
     (1,"Notificación de solicitud urgente.",
      "En caso de publicarse una solicitud urgente con tipo de sangre compatible y dentro de la zona del usuario.",
      "El sistema detecta la nueva solicitud.",
      "El sistema envía una notificación push: '¡Solicitud urgente! Sangre [tipo] necesaria a [distancia] km de ti'."),
     (2,"Notificación de donante aceptado.",
      "En caso de que un donante confirme su participación en la solicitud del usuario.",
      "El donante confirma la donación.",
      "El sistema notifica al solicitante: '[Nombre del donante] ha aceptado donar [tipo de sangre] — acudirá a [centro de salud]'."),
     (3,"Notificación de recordatorio de campaña.",
      "En caso de que el usuario tenga un cupo reservado en una campaña que se realizará en las próximas 24 horas.",
      "El sistema detecta la proximidad de la campaña.",
      "El sistema envía la notificación push: 'Recuerda: mañana tienes reservado un cupo en la campaña [nombre] en [lugar] a las [hora]'."),
     (4,"Notificación de reconocimiento o insignia.",
      "En caso de que el usuario alcance un hito (primera donación, 5 donaciones, etc.).",
      "El sistema registra el hito al confirmar la donación.",
      "El sistema envía la notificación push: '¡Felicidades! Has desbloqueado la insignia [nombre]' y actualiza el perfil."),
     (5,"Ver centro de notificaciones.",
      "En caso de tener notificaciones pendientes.",
      "Clic en el ícono de notificaciones en el Home.",
      "El sistema muestra el historial de notificaciones diferenciadas visualmente: no leídas (fondo blanco con punto rojo) y leídas (fondo gris)."),
     (6,"Marcar todas como leídas.",
      "En caso de tener notificaciones no leídas.",
      "Clic en la opción MARCAR TODO COMO LEÍDO.",
      "El sistema actualiza el estado de todas las notificaciones a 'leída' y elimina el indicador numérico en el ícono."),
   ]
  ),
],

# ══════════════════════════════════════════════════════════════
"HE-08": [
  ("HE-08","HU-8","Yo como USUARIO AUTENTICADO",
   "Necesito que la aplicación proteja mis datos personales.",
   "Para garantizar la privacidad de mi información sensible dentro de la plataforma.",
   [
     (1,"Verificación de identidad con foto frontal, trasera y selfie.",
      "En caso de completar el Paso 2 del registro con foto frontal del documento, foto trasera y selfie obligatoria con cámara.",
      "El sistema procesa las tres imágenes cargadas.",
      "El sistema otorga el sello 'Verificado' al perfil del usuario, visible para otros usuarios en solicitudes y campañas."),
     (2,"Acceso a configuración de privacidad.",
      "En caso de estar en la pantalla de perfil.",
      "Clic en la opción PRIVACIDAD.",
      "El sistema muestra la pantalla de configuración de privacidad con los controles de visibilidad del perfil."),
     (3,"Datos de contacto no expuestos públicamente.",
      "En caso de que otro usuario visualice el perfil o resumen de un donante/solicitante.",
      "El usuario accede a un perfil ajeno en solicitudes o campañas.",
      "El sistema no muestra el número de teléfono, correo electrónico ni número de cédula en ninguna vista pública."),
   ]
  ),
  ("HE-08","HU-8.1","Yo como ADMINISTRADOR DE LA PLATAFORMA",
   "Necesito que el sistema prevenga el uso indebido de funciones críticas.",
   "Para mantener la integridad y confianza en la plataforma !Blood.",
   [
     (1,"Detección de abuso de urgencia.",
      "En caso de que un mismo usuario cree repetidamente solicitudes de urgencia 'Crítica' sin completarlas.",
      "El sistema analiza el historial de solicitudes del usuario.",
      "El sistema activa el proceso de penalización de HE-11 y restringe temporalmente la creación de solicitudes urgentes."),
     (2,"Reporte de calificación falsa.",
      "En caso de que una calificación sea reportada por el receptor como falsa o malintencionada.",
      "El receptor clic en REPORTAR CALIFICACIÓN.",
      "El sistema registra el reporte, notifica al administrador para revisión y aplica la penalización correspondiente tras validación (HE-11)."),
   ]
  ),
],

# ══════════════════════════════════════════════════════════════
"HE-09": [
  ("HE-09","HU-9","Yo como USUARIO AUTENTICADO",
   "Necesito ver estadísticas globales de la comunidad !Blood y mi reputación personal.",
   "Para sentirme parte de una red con impacto real, mantenerme motivado a donar y conocer mi calificación promedio.",
   [
     (1,"Ver estadísticas en Home.",
      "En caso de estar en la pantalla principal.",
      "El usuario visualiza la sección de estadísticas de comunidad.",
      "El sistema muestra el total de donantes activos y el total de vidas salvadas en la plataforma, actualizados en tiempo real."),
     (2,"Ver calificación promedio en perfil.",
      "En caso de haber recibido al menos una calificación de la comunidad.",
      "El usuario navega a la sección PERFIL.",
      "El sistema muestra las estrellas con el promedio numérico (ej. 4.8 ★) bajo el nombre y tipo de sangre del usuario."),
     (3,"Recibir reconocimiento por logro.",
      "En caso de completar una donación o alcanzar un hito (ej. 5 donaciones).",
      "El sistema registra la donación o el hito.",
      "El sistema envía una notificación de reconocimiento y muestra la insignia correspondiente en el perfil del usuario."),
     (4,"Ver insignias en perfil.",
      "En caso de haber obtenido al menos una insignia.",
      "El usuario navega a la pantalla de Perfil.",
      "El sistema muestra las insignias obtenidas (Verificado, Donante Activo, Héroe de la Comunidad, etc.) de forma destacada bajo el nombre."),
   ]
  ),
],

# ══════════════════════════════════════════════════════════════
"HE-10": [
  ("HE-10","HU-10","Yo como SISTEMA",
   "Necesito calcular automáticamente la compatibilidad entre tipos de sangre.",
   "Para conectar donantes compatibles con las solicitudes y campañas correctas.",
   [
     (1,"Compatibilidad correcta A+.",
      "En caso de un donante con tipo A+.",
      "El sistema evalúa compatibilidad al mostrar solicitudes.",
      "El sistema identifica como compatibles las solicitudes de tipo A+ y AB+."),
     (2,"Compatibilidad O- universal.",
      "En caso de un donante con tipo O- (donante universal).",
      "El sistema evalúa compatibilidad al mostrar solicitudes.",
      "El sistema identifica como compatibles las solicitudes de TODOS los tipos de sangre y aplica el chip 'Compatible' a todas."),
     (3,"Incompatibilidad notificada.",
      "En caso de que el tipo de sangre del donante no sea compatible con el requerido.",
      "El usuario abre una solicitud incompatible.",
      "El sistema muestra el aviso de incompatibilidad y desactiva el botón QUIERO DONAR."),
   ]
  ),
  ("HE-10","HU-10.1","Yo como SISTEMA",
   "Necesito calcular la distancia entre el usuario y cada solicitud o campaña.",
   "Para mostrar información georreferenciada relevante y ordenar resultados por cercanía.",
   [
     (1,"Distancia calculada en listado.",
      "En caso de que el usuario tenga su ubicación activada.",
      "El usuario accede al listado de solicitudes o campañas.",
      "El sistema muestra la distancia en km o metros para cada ítem del listado, ordenados de más cercano a más lejano por defecto."),
     (2,"Permiso de ubicación denegado.",
      "En caso de que el usuario haya denegado el permiso de geolocalización.",
      "El sistema intenta calcular la distancia.",
      "El sistema muestra el mensaje: 'Activa tu ubicación para ver solicitudes y campañas cercanas a ti' con el botón ACTIVAR UBICACIÓN."),
   ]
  ),
],

# ══════════════════════════════════════════════════════════════
"HE-11": [
  ("HE-11","HU-11","Yo como SISTEMA",
   "Necesito aplicar penalizaciones automáticas ante conductas indebidas.",
   "Para mantener la integridad de la plataforma y proteger a los usuarios de abusos.",
   [
     (1,"Penalización por abuso de urgencia.",
      "En caso de que un usuario haya creado más de 3 solicitudes 'Críticas' en 7 días sin que ninguna sea completada.",
      "El sistema detecta el patrón de abuso.",
      "El sistema bloquea temporalmente al usuario para crear nuevas solicitudes urgentes, le notifica la restricción y registra el incidente en el log de moderación."),
     (2,"Penalización por cancelación de donación confirmada.",
      "En caso de que un donante cancele una donación que ya había confirmado.",
      "El donante cancela la donación tras la confirmación.",
      "El sistema aplica la penalización de 30 días: el donante no puede confirmar nuevas donaciones durante ese período y ve el aviso: 'Tu cuenta tiene una restricción activa por 30 días debido a cancelación de donación confirmada'."),
     (3,"Penalización por calificación discriminatoria.",
      "En caso de que se detecte o reporte una calificación con lenguaje discriminatorio, ofensivo o malintencionado.",
      "El sistema o el administrador validan el reporte.",
      "El sistema suspende la cuenta del usuario infractor, le notifica la suspensión con el motivo y registra el caso para revisión del equipo de moderación."),
     (4,"Consultar estado de penalización.",
      "En caso de que el usuario desee saber si tiene alguna penalización activa.",
      "El usuario navega a la sección PRIVACIDAD o CONFIGURACIÓN en su perfil.",
      "El sistema muestra el estado de sanciones activas con: tipo de penalización, motivo, fecha de inicio y fecha de expiración (si aplica)."),
   ]
  ),
],

# ══════════════════════════════════════════════════════════════
"HE-12": [
  ("HE-12","HU-12","Yo como USUARIO AUTENTICADO",
   "Necesito completar la encuesta de aptitud para donación.",
   "Para saber si estoy en condiciones de donar sangre de forma segura y responsable.",
   [
     (1,"Encuesta presentada tras registro exitoso.",
      "En caso de haber completado el Paso 3 del registro correctamente.",
      "El sistema redirige al usuario tras la pantalla de bienvenida.",
      "El sistema muestra automáticamente la pantalla de Encuesta de Aptitud con 8 preguntas de respuesta Sí/No "
      "sobre condiciones de salud, medicamentos, viajes recientes y comportamientos de riesgo."),
     (2,"Resultado: Apto para donar.",
      "En caso de que el usuario responda las 8 preguntas sin indicadores de exclusión.",
      "El usuario completa la encuesta y presiona FINALIZAR.",
      "El sistema muestra el mensaje: '¡Excelente! Estás en condiciones de donar sangre. Gracias por tu compromiso con la vida.' "
      "y actualiza el perfil con el estado 'Apto para donar'."),
     (3,"Resultado: No apto temporalmente.",
      "En caso de que alguna respuesta indique una condición de exclusión temporal (ej. medicamento, viaje reciente).",
      "El usuario completa la encuesta.",
      "El sistema muestra el mensaje: 'Por el momento no puedes donar sangre. Te recomendamos consultar con un profesional de salud. "
      "Podrás retomar la encuesta en 30 días.' y actualiza el estado del perfil."),
     (4,"Omitir encuesta por ahora.",
      "En caso de que el usuario no desee completar la encuesta en ese momento.",
      "Clic en la opción COMPLETAR MÁS TARDE.",
      "El sistema guarda el estado como 'Encuesta pendiente' y redirige al Home. El estado aparece visible en el perfil."),
     (5,"Retomar encuesta desde perfil.",
      "En caso de tener la encuesta pendiente o querer actualizarla.",
      "Clic en la opción ENCUESTA DE APTITUD en el menú del perfil.",
      "El sistema carga la pantalla de encuesta con los campos en blanco o con las respuestas previas para revisión."),
     (6,"Pregunta no respondida.",
      "En caso de intentar finalizar la encuesta con al menos una pregunta sin responder.",
      "Clic en FINALIZAR.",
      "El sistema resalta las preguntas pendientes y muestra el mensaje: 'Por favor responde todas las preguntas para continuar'."),
   ]
  ),
],

}

# ──────────────────────────────────────────────────────────────
# CONSTRUCCIÓN DEL WORKBOOK
# ──────────────────────────────────────────────────────────────

wb = Workbook()
wb.remove(wb.active)   # quitar la hoja por defecto

# ──────────────────────────────────────
# HOJA 1: HISTORIAS ÉPICAS
# ──────────────────────────────────────
ws_ep = wb.create_sheet("Historias Épicas")

# Fila 1 – vacía con altura grande (simula la imagen)
ws_ep.row_dimensions[1].height = 18
ws_ep.merge_cells("A1:C1")

# Fila 2 – Título principal
ws_ep.merge_cells("A2:C2")
ws_ep.row_dimensions[2].height = 30
title_cell = ws_ep["A2"]
title_cell.value = "HISTORIAS EPICAS"
title_cell.font = Font(bold=True, size=14, color=BLACK, name="Calibri")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
title_cell.fill = fill("D9D9D9")
title_cell.border = border_all()

# Fila 3 vacía
ws_ep.row_dimensions[3].height = 14
ws_ep.merge_cells("A3:C3")
ws_ep["A3"].fill = fill("FFFFFF")
ws_ep["A3"].border = border_all()

# Fila 4 – Encabezados de tabla
headers_ep = ["HE ID", "TÍTULO", "DESCRIPCIÓN"]
ws_ep.row_dimensions[4].height = 32
for col, h in enumerate(headers_ep, 1):
    c = ws_ep.cell(row=4, column=col, value=h)
    c.fill = fill(DARK_HEADER)
    c.font = font(bold=True, color=WHITE, size=11)
    c.alignment = center_wrap()
    c.border = border_all()

# Ancho de columnas
ws_ep.column_dimensions["A"].width = 10
ws_ep.column_dimensions["B"].width = 38
ws_ep.column_dimensions["C"].width = 70

# Filas de datos
for i, (he_id, titulo, desc) in enumerate(epicas, 5):
    ws_ep.row_dimensions[i].height = 42
    bg = "FFFFFF" if i % 2 == 1 else ROW_ALT

    c_id = ws_ep.cell(row=i, column=1, value=he_id)
    c_id.fill = fill(bg)
    c_id.font = font(bold=True, color=ORANGE_LINK, size=10)
    c_id.alignment = center_wrap()
    c_id.border = border_all()

    c_ti = ws_ep.cell(row=i, column=2, value=titulo)
    c_ti.fill = fill(bg)
    c_ti.font = font(color=ORANGE_LINK, size=10)
    c_ti.alignment = center_wrap()
    c_ti.border = border_all()

    c_de = ws_ep.cell(row=i, column=3, value=desc)
    c_de.fill = fill(bg)
    c_de.font = font(color=ORANGE_LINK, size=10)
    c_de.alignment = left_wrap()
    c_de.border = border_all()

# ──────────────────────────────────────
# HOJAS POR ÉPICA
# ──────────────────────────────────────
COLS_HU = ["A","B","C","D","E","F","G","H","I","J","K"]
COL_WIDTHS = [12, 12, 18, 26, 22, 10, 22, 32, 28, 40]  # A-J

SHEET_NAMES = {
    "HE-01": "HE-01 Acceso Usuarios",
    "HE-02": "HE-02 Perfil Usuario",
    "HE-03": "HE-03 Solicitudes Sangre",
    "HE-04": "HE-04 Proceso Donacion",
    "HE-05": "HE-05 Calificacion Mutua",
    "HE-06": "HE-06 Campanas",
    "HE-07": "HE-07 Notificaciones",
    "HE-08": "HE-08 Moderacion Seguridad",
    "HE-09": "HE-09 Estadisticas",
    "HE-10": "HE-10 Compat Geolocalizacion",
    "HE-11": "HE-11 Penalizaciones",
    "HE-12": "HE-12 Encuesta Aptitud",
}

EPICA_TITLES = {
    "HE-01": "Historias de USUARIO y criterios de aceptación de HE-1",
    "HE-02": "Historias de USUARIO y criterios de aceptación de HE-2",
    "HE-03": "Historias de USUARIO y criterios de aceptación de HE-3",
    "HE-04": "Historias de USUARIO y criterios de aceptación de HE-4",
    "HE-05": "Historias de USUARIO y criterios de aceptación de HE-5",
    "HE-06": "Historias de USUARIO y criterios de aceptación de HE-6",
    "HE-07": "Historias de USUARIO y criterios de aceptación de HE-7",
    "HE-08": "Historias de USUARIO y criterios de aceptación de HE-8",
    "HE-09": "Historias de USUARIO y criterios de aceptación de HE-9",
    "HE-10": "Historias de USUARIO y criterios de aceptación de HE-10",
    "HE-11": "Historias de USUARIO y criterios de aceptación de HE-11",
    "HE-12": "Historias de USUARIO y criterios de aceptación de HE-12",
}

for epica_id, historias in hu_por_epica.items():
    ws = wb.create_sheet(SHEET_NAMES[epica_id])

    # ── Anchos de columna ──
    for col_i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = w

    # ── Fila 1: Título ──
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 38
    t = ws["A1"]
    t.value = EPICA_TITLES[epica_id]
    t.font = Font(bold=True, size=14, color=BLACK, name="Calibri")
    t.alignment = Alignment(horizontal="center", vertical="center")

    # ── Fila 2 vacía ──
    ws.merge_cells("A2:J2")
    ws.row_dimensions[2].height = 10

    # ── Fila 3: Grupo de encabezados "Enunciado" y "Criterios" ──
    ws.row_dimensions[3].height = 22
    ws.merge_cells("A3:E3")  # Enunciado (cols A-E / 1-5 … pero ver abajo)
    # Ajustar: A=ID épica, B=ID hist, C=Rol, D=Quiero, E=Para → enunciado=A-E
    # F=Número escenario, G=Título criterio, H=Contexto, I=Evento, J=Resultado → criterios=F-J
    ws.merge_cells("F3:J3")

    ce_e = ws["A3"]
    ce_e.value = "Enunciado de la historia"
    ce_e.fill = fill(MID_BLUE)
    ce_e.font = font(bold=True, color=WHITE, size=10)
    ce_e.alignment = center_wrap()
    ce_e.border = border_all()
    for col in ["B3","C3","D3","E3"]:
        ws[col].fill = fill(MID_BLUE)
        ws[col].border = border_all()

    ce_c = ws["F3"]
    ce_c.value = "Criterios de aceptación"
    ce_c.fill = fill(MID_BLUE)
    ce_c.font = font(bold=True, color=WHITE, size=10)
    ce_c.alignment = center_wrap()
    ce_c.border = border_all()
    for col in ["G3","H3","I3","J3"]:
        ws[col].fill = fill(MID_BLUE)
        ws[col].border = border_all()

    # ── Fila 4: Sub-encabezados ──
    ws.row_dimensions[4].height = 46
    sub_headers = [
        ("A4","Identificador\n(ID) de la\népica"),
        ("B4","Identificador\n(ID) de la\nhistoria"),
        ("C4","Rol"),
        ("D4","Característica /\nFuncionalidad\n(Quiero)"),
        ("E4","Razón /\nResultado\n(Para)"),
        ("F4","Número (#)\nde\nescenario"),
        ("G4","Criterio de\naceptación\n(Título)"),
        ("H4","Contexto"),
        ("I4","Evento"),
        ("J4","Resultado /\nComportamiento\nesperado"),
    ]
    for addr, label in sub_headers:
        c = ws[addr]
        c.value = label
        c.fill = fill(BLUE_HEADER)
        c.font = font(bold=True, color=WHITE, size=9)
        c.alignment = center_wrap()
        c.border = border_all()

    # ── Filas de datos ──
    current_row = 5

    for (id_ep, id_hu, rol, quiero, para, escenarios) in historias:
        n_esc = len(escenarios)
        end_row = current_row + n_esc - 1

        # Merge celdas A-E para los campos del enunciado
        for col_letter in ["A","B","C","D","E"]:
            if n_esc > 1:
                ws.merge_cells(f"{col_letter}{current_row}:{col_letter}{end_row}")

        # Valores del enunciado
        row_data = {
            "A": (id_ep,   True,  ORANGE_LINK),
            "B": (id_hu,   True,  ORANGE_LINK),
            "C": (rol,     False, BLACK),
            "D": (quiero,  False, BLACK),
            "E": (para,    False, BLACK),
        }
        for col_letter, (val, bold, color) in row_data.items():
            c = ws[f"{col_letter}{current_row}"]
            c.value = val
            c.font = font(bold=bold, color=color, size=9)
            c.alignment = center_wrap()
            c.border = border_all()

        # Escenarios
        alt = True
        for esc_row, (num, tit, ctx, evt, res) in enumerate(escenarios):
            r = current_row + esc_row
            ws.row_dimensions[r].height = 54
            bg = "FFFFFF" if alt else ROW_ALT
            alt = not alt

            esc_data = [
                ("F", str(num), True,  BLACK),
                ("G", tit,      False, BLACK),
                ("H", ctx,      False, BLACK),
                ("I", evt,      False, BLACK),
                ("J", res,      False, BLACK),
            ]
            for col_letter, val, bold, color in esc_data:
                c = ws[f"{col_letter}{r}"]
                c.value = val
                c.font = font(bold=bold, color=color, size=9)
                c.alignment = left_wrap()
                c.border = border_all()

            # Fondo para celdas A-E de las filas adicionales (merge ya aplicado)
            for col_letter in ["A","B","C","D","E"]:
                ws[f"{col_letter}{r}"].fill = fill(bg)

        current_row = end_row + 1

# ──────────────────────────────────────────────────────────────
# GUARDAR
# ──────────────────────────────────────────────────────────────
output_path = "Historias_de_Usuario_Blood_v3.xlsx"
wb.save(output_path)
print(f"Archivo generado: {output_path}")
